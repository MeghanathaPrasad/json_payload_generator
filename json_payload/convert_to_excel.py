# import pandas as pd
#
# # Your JSON data
# data = {
#     "accountvault": {
#         "title": "Test CC Account",
#         "payment_method": "cc",
#         "account_holder_name": "John Smith",
#         "account_number": "5454545454545454",
#         "exp_date": "0919",
#         "contact_id": "54c285f1-678d-9510-318c-bcb13ab0c1c2",
#         "location_id": "sd9fusd9f-678d-9510-318c-bcb13ab4324"
#     }
# }
#
# # Convert JSON data to a DataFrame
# df = pd.DataFrame([data['accountvault']])
#
# # Write the DataFrame to an Excel file
# df.to_excel('accountvault_data.xlsx', index=False)
#
# print("Data successfully written to Excel.")




#
# import pandas as pd
#
# # Your JSON data
# data = {
#     "accountvault": {
#         "title": "Test CC Account",
#         "payment_method": "cc",
#         "account_holder_name": "John Smith",
#         "account_number": "5454545454545454",
#         "exp_date": "0919",
#         "contact_id": "54c285f1-678d-9510-318c-bcb13ab0c1c2",
#         "location_id": "sd9fusd9f-678d-9510-318c-bcb13ab4324"
#     }
# }
#
# # Create a new dict with "accountvault" as a prefix
# modified_data = {f"accountvault_{k}": v for k, v in data['accountvault'].items()}
#
# # Convert JSON data to a DataFrame
# df = pd.DataFrame([modified_data])
#
# # Write the DataFrame to an Excel file with a custom sheet name
# with pd.ExcelWriter('accountvault_data.xlsx', engine='xlsxwriter') as writer:
#     df.to_excel(writer, sheet_name='create_accountvault', index=False)
#
# print("Data successfully written to Excel with sheet name 'create_accountvault'.")




# import pandas as pd
#
# # Your JSON data
# data = {
#     "accountvault": {
#         "title": "Test CC Account",
#         "payment_method": "cc",
#         "account_holder_name": "John Smith",
#         "account_number": "5454545454545454",
#         "exp_date": "0919",
#         "contact_id": "54c285f1-678d-9510-318c-bcb13ab0c1c2",
#         "location_id": "sd9fusd9f-678d-9510-318c-bcb13ab4324"
#     }
# }
#
# # Create a DataFrame for the actual data
# df = pd.DataFrame([data['accountvault']])
#
# # Write the DataFrame to an Excel file with a custom sheet name
# with pd.ExcelWriter('accountvault_data.xlsx', engine='xlsxwriter') as writer:
#     # Create the Excel sheet with the desired sheet name
#     sheet_name = 'create_accountvault'
#
#     # Write the table name ("accountvault") as a header row
#     df_table_name = pd.DataFrame([["accountvault"]])
#     df_table_name.to_excel(writer, sheet_name=sheet_name, header=False, index=False, startrow=0)
#
#     # Write the actual data below the table name
#     df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
#
# print("Data successfully written to Excel with 'accountvault' as the table name.")

#
# {
#         "title": "Test ACH Account",
#         "payment_method": "ach",
#         "account_holder_name": "John Smith",
#         "account_number":"700953657",
#         "account_type": "checking",
#         "is_company": "false",
#         "routing": "100020200",
#         "contact_id": "54c285f1-678d-9510-318c-bcb13ab0c1c2",
#         "location_id": "sd9fusd9f-678d-9510-318c-bcb13ab4324"
#     }


# import os
# import pandas as pd
# from openpyxl import load_workbook
#
#
# def add_sheet_to_excel(file_name, sheet_name, data, table_name):
#     if os.path.exists(file_name):
#         try:
#             # Open the existing workbook
#             book = load_workbook(file_name)
#
#             # Use ExcelWriter with append mode
#             with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
#                 # Create DataFrame for the new data and table name
#                 # First row is the table name, second row onwards is the data
#                 table_name_df = pd.DataFrame([[table_name]])  # Create table name as DataFrame
#                 data_df = pd.DataFrame([data])  # Create data as DataFrame
#
#                 # Write both table name and data in one operation
#                 table_name_df.to_excel(writer, sheet_name=sheet_name, header=False, index=False, startrow=0)
#                 data_df.to_excel(writer, sheet_name=sheet_name, header=True, index=False, startrow=1)
#
#                 print(f"Data successfully added to '{sheet_name}' in {file_name}.")
#         except Exception as e:
#             print(f"Error reading the file: {e}. It may be corrupted or in use.")
#     else:
#         # Create a new file and add the first sheet
#         with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
#             table_name_df = pd.DataFrame([[table_name]])  # Create table name as DataFrame
#             data_df = pd.DataFrame([data])  # Create data as DataFrame
#
#             # Write both table name and data in one operation
#             table_name_df.to_excel(writer, sheet_name=sheet_name, header=False, index=False, startrow=0)
#             data_df.to_excel(writer, sheet_name=sheet_name, header=True, index=False, startrow=1)
#             print(f"File created and data written to '{sheet_name}'.")
#
#
# # Example usage
# data = {
#         "title": "Test ACH Account",
#         "payment_method": "ach",
#         "account_holder_name": "John Smith",
#         "account_number":"700953657",
#         "account_type": "checking",
#         "is_company": "false",
#         "routing": "100020200",
#         "contact_id": "54c285f1-678d-9510-318c-bcb13ab0c1c2",
#         "location_id": "sd9fusd9f-678d-9510-318c-bcb13ab4324"
#     }
#
# file_name = 'accountvault_data.xlsx'
# sheet_name = input("Enter sheet name: ")  # Example: 'Create Bank Account Vault(ACH)'
# table_name = input("Enter table name: ")  # Example: 'accountvault'
#
# # Add sheet to the existing Excel file
# add_sheet_to_excel(file_name, sheet_name, data, table_name)
#

#working
# import os
# import pandas as pd
# from openpyxl import load_workbook
#
#
# def add_sheet_to_excel(file_name, sheet_name, data, table_name):
#     if os.path.exists(file_name):
#         try:
#             # Load the existing workbook
#             book = load_workbook(file_name)
#
#             # Check if the sheet already exists
#             if sheet_name in book.sheetnames:
#                 print(f"Sheet '{sheet_name}' already exists. Please choose a different name.")
#                 return
#
#             # Use ExcelWriter with append mode
#             with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
#                 writer.book = book
#
#                 # Create DataFrame for the new data
#                 df = pd.DataFrame([data])
#
#                 # Write table name at the top of the new sheet
#                 df_table_name = pd.DataFrame([[table_name]])
#                 df_table_name.to_excel(writer, sheet_name=sheet_name, header=False, index=False, startrow=0)
#
#                 # Write data below the table name
#                 df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
#
#                 print(f"Data successfully added to '{sheet_name}' in {file_name}.")
#         except Exception as e:
#             print(f"Error reading the file: {e}. It may be corrupted or in use.")
#     else:
#         # Create a new file and add the first sheet
#         with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
#             df = pd.DataFrame([data])
#             df_table_name = pd.DataFrame([[table_name]])
#             df_table_name.to_excel(writer, sheet_name=sheet_name, header=False, index=False, startrow=0)
#             df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
#             print(f"File created and data written to '{sheet_name}'.")
#
#
# # Example usage:
# data = {
#         "title": "Test ACH Account",
#         "payment_method": "ach",
#         "account_holder_name": "John Smith",
#         "account_number":"700953657",
#         "account_type": "checking",
#         "is_company": "false",
#         "routing": "100020200",
#         "contact_id": "54c285f1-678d-9510-318c-bcb13ab0c1c2",
#         "location_id": "sd9fusd9f-678d-9510-318c-bcb13ab4324"
#     }
#
# file_name = 'accountvault_data.xlsx'
# sheet_name = input("Enter sheet name: ")  # Example: 'Create Bank Account Vault(ACH)'
# table_name = input("Enter table name: ")  # Example: 'accountvault'
#
# # Add sheet to the existing Excel file
# add_sheet_to_excel(file_name, sheet_name, data, table_name)

# import os
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
#
#
# def add_sheet_to_excel(file_name, sheet_name, data, table_name):
#     data_df = pd.DataFrame([data])
#     table_name_df = pd.DataFrame([[table_name]])
#
#     if os.path.exists(file_name):
#         try:
#             # Load the existing workbook
#             book = load_workbook(file_name)
#
#             # Check if the sheet already exists
#             if sheet_name in book.sheetnames:
#                 # Append to the existing sheet
#                 sheet = book[sheet_name]
#
#                 # Write table name at the top of the existing sheet
#                 for r in dataframe_to_rows(table_name_df, index=False, header=False):
#                     sheet.append(r)
#                 for r in dataframe_to_rows(data_df, index=False, header=False):
#                     sheet.append(r)
#
#             else:
#                 # Create a new sheet
#                 sheet = book.create_sheet(title=sheet_name)
#                 for r in dataframe_to_rows(table_name_df, index=False, header=False):
#                     sheet.append(r)
#                 for r in dataframe_to_rows(data_df, index=False, header=False):
#                     sheet.append(r)
#
#             # Save the workbook
#             book.save(file_name)
#             print(f"Data successfully added to '{sheet_name}' in {file_name}.")
#
#         except Exception as e:
#             print(f"Error processing the file: {e}.")
#     else:
#         # Create a new file and add the first sheet
#         with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
#             # Create the DataFrame and write the table name and data
#             table_name_df.to_excel(writer, sheet_name=sheet_name, header=False, index=False, startrow=0)
#             data_df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
#
#         print(f"File created and data written to '{sheet_name}'.")
#
#
# # Example usage
# data = {
#     "title": "Test ACH Account",
#     "payment_method": "ach",
#     "account_holder_name": "John Smith",
#     "account_number": "700953657",
#     "account_type": "checking",
#     "is_company": "false",
#     "routing": "100020200",
#     "contact_id": "54c285f1-678d-9510-318c-bcb13ab0c1c2",
#     "location_id": "sd9fusd9f-678d-9510-318c-bcb13ab4324"
# }
#
# file_name = 'accountvault_data.xlsx'
# sheet_name = input("Enter sheet name: ")
# table_name = input("Enter table name: ")
#
# # Add sheet to the existing Excel file
# add_sheet_to_excel(file_name, sheet_name, data, table_name)


import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def add_sheet_to_excel(file_name, sheet_name, data, table_name):
    data_df = pd.DataFrame([data])
    table_name_df = pd.DataFrame([[table_name]])

    if os.path.exists(file_name):
        try:
            # Load the existing workbook
            book = load_workbook(file_name)

            # Check if the sheet already exists
            if sheet_name in book.sheetnames:
                # Append to the existing sheet
                sheet = book[sheet_name]

                # Find the next empty row
                next_row = sheet.max_row + 1

                # Write table name and data at the next empty row
                for r in dataframe_to_rows(table_name_df, index=False, header=False):
                    sheet.append(r)
                for r in dataframe_to_rows(data_df, index=False, header=False):
                    sheet.append(r)

            else:
                # Create a new sheet
                sheet = book.create_sheet(title=sheet_name)


                # Write table name and data
                for r in dataframe_to_rows(table_name_df, index=False, header=False):
                    sheet.append(r)
                for r in dataframe_to_rows(data_df, index=False, header=True):
                    sheet.append(r)




            # Save the workbook
            book.save(file_name)
            print(f"Data successfully added to '{sheet_name}' in {file_name}.")

        except Exception as e:
            print(f"Error processing the file: {e}.")
    else:
        # Create a new file and add the first sheet
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            # Create the DataFrame and write the table name and data
            table_name_df.to_excel(writer, sheet_name=sheet_name, header=False, index=False, startrow=0)
            data_df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)

        print(f"File created and data written to '{sheet_name}'.")


# Example usage
data = {
        "title": "Test CC Account",
        "payment_method": "cc",
        "account_holder_name": "John Smith",
        "account_number":"5454545454545454",
        "exp_date": "0919",
        "contact_id": "54c285f1-678d-9510-318c-bcb13ab0c1c2",
        "location_id": "sd9fusd9f-678d-9510-318c-bcb13ab4324"
    }
file_name = 'accountvault_data.xlsx'
sheet_name = input("Enter sheet name: ")
table_name = input("Enter table name: ")

# Add sheet to the existing Excel file
add_sheet_to_excel(file_name, sheet_name, data, table_name)



