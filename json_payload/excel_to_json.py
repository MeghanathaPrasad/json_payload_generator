# import json
# import pandas as pd
#
# with open('data.json', 'r') as file:
#     nested_dict = json.load(file)
#
# def flatten_dict(d, parent_key='', sep='.'):
#     """
#     Recursively flattens a nested dictionary and handles lists, tuples, and other data types.
#
#     Args:
#     - d (dict): The nested dictionary to flatten.
#     - parent_key (str): The base key string for the flattened keys.
#     - sep (str): The separator between keys.
#
#     Returns:
#     - dict: The flattened dictionary.
#     """
#     flat_dict = {}
#     for k, v in d.items():
#         new_key = f'{parent_key}{sep}{k}' if parent_key else k  # Construct new key
#
#         if isinstance(v, dict):  # Handle nested dictionary
#             flat_dict.update(flatten_dict(v, new_key, sep=sep))
#
#         elif isinstance(v, (list, tuple)):  # Handle lists or tuples
#             for i, item in enumerate(v):
#                 flat_dict.update(flatten_dict({f"{new_key}{sep}{i}": item}, '', sep=sep))
#
#         else:  # Handle other types (str, int, float, etc.)
#             flat_dict[new_key] = v
#
#     return flat_dict
#
#
# # Flatten the nested dictionary
# flattened_dict = flatten_dict(nested_dict)
#
# # print(flattened_dict)flattened_dict
#
# # Convert flattened dictionary to a list of key-value pairs
# data = list(flattened_dict.items())
#
# # Convert to DataFrame
# df = pd.DataFrame(data, columns=['Key', 'Value'])
#
# print(df)
#
# df.to_excel("output.xlsx", index=False)
#


#
# import pandas as pd
# import json
#
# # Load Excel data into a DataFrame
# df = pd.read_excel('output.xlsx', sheet_name='Sheet1')  # Replace 'your_file.xlsx' with your file path
#
# # Convert DataFrame to a dictionary
# data_dict = df.to_dict(orient='records')
#
# print(data_dict)
#
# nested_dict = {}
#
#
# for d in data_dict:
#     a = 0
#     keys = d.get('Key')
#     keys_list = keys.split('.')
#     # print(keys_list)
#
#     data = nested_dict
#     length = len(keys_list)
#
#     for i in range(length):
#         key = keys_list[i]
#         if not data.get(key):
#             if i != length-1:
#                 data.update({key: {}})
#             else:
#                 data.update({key: d.get('Value')})
#
#         data = data.get(key)
#
# print(nested_dict)
#
# # Save JSON data to a file
# with open('result.json', 'w') as file:
#     json.dump(nested_dict, file, indent=4)







import pandas as pd
import json
from openpyxl import load_workbook

workbook = load_workbook("output.xlsx")

# Get the sheet by name (replace 'nested data' with your actual sheet name)
sheet = workbook["Sheet2"]

# Find the first non-empty cell in the sheet
for row in sheet.iter_rows(values_only=True):
    for cell in row:
        if cell is not None:  # Checking for the first non-empty cell
            first_word = str(cell)
            break
    if cell is not None:
        break

# Load Excel data into a DataFrame
df = pd.read_excel('output.xlsx', sheet_name='Sheet2',skiprows=1)  # Replace 'your_file.xlsx' with your file path
# df = df.iloc[2:, :2]
# Convert DataFrame to a dictionary
data_dict = df.to_dict(orient='records')

# print(data_dict)

nested_dict = {}


for d in data_dict:
    keys = d.get('Key')
    keys_list = keys.split('.')
    length = len(keys_list)
    # print(keys_list)

    current = nested_dict
    for i, key in enumerate(keys_list):
        val = None
        try:
            if keys_list[i+1].isdecimal():
                val = []
            else:
                val = {}
        except:
            pass

        if isinstance(current, dict):
            if not current.get(key):
                if i != length-1:
                    current.update({key: val})
                else:
                    current.update({key: d.get('Value')})
            current = current.get(key)

        elif isinstance(current, list):
            key = int(key)
            if len(current[:key+1]) != key+1:
                if i != length-1:
                    current.insert(key, {})
                else:
                    current.insert(key, d.get('Value'))
            current = current[key]


# print(nested_dict)
result = {
    first_word: nested_dict
}
print(result)
# Save JSON data to a file
with open('result.json', 'w') as file:
    json.dump(result, file, indent=4)


