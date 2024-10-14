# import pandas as pd
# from collections import defaultdict
#
#
# def unflatten_json(flattened_dict, sep='_'):
#     """
#     Convert flattened dictionary back to nested dictionary.
#     """
#     nested_dict = defaultdict(dict)
#     for key, value in flattened_dict.items():
#         parts = key.split(sep)
#         d = nested_dict
#         for part in parts[:-1]:
#             if part.isdigit():
#                 part = int(part)
#             d = d.setdefault(part, {})
#         d[parts[-1]] = value
#     return dict(nested_dict)
#
# def excel_to_json(file_name, sheet_name, start_row=2):
#     """
#     Read Excel file and convert key-value pairs starting from a specific row to JSON format.
#     """
#     # Read the Excel sheet
#     df = pd.read_excel(file_name, sheet_name=sheet_name, header=None)
#
#     # Print the columns to debug
#     print("Columns in the Excel sheet:", df.columns)
#
#     # Skip rows until you reach the key-value pairs (starting from the provided start_row)
#     df = df.iloc[start_row:, :2]  # Adjust to take only key-value pairs (2 columns)
#     df.columns = ['Key', 'Value']  # Rename the columns after slicing
#
#     # Convert the dataframe to dictionary
#     flattened_dict = dict(zip(df['Key'], df['Value']))
#
#     # Convert the flattened dictionary back to nested structure
#     nested_data = unflatten_json(flattened_dict)
#
#     return nested_data
#
# # Example usage
# file_name = 'accountvault_data.xlsx'
# sheet_name = input("Enter sheet name: ")
#
# # Convert the Excel back to JSON
# nested_json = excel_to_json(file_name, sheet_name, start_row=2)  # Adjust 'start_row' if key-value pairs start later
#
# print(nested_json)

#
# import pandas as pd
# import json
#
#
# def unflatten_json(flattened_dict, sep='_'):
#     """
#     Convert flattened dictionary back to nested dictionary.
#     Handles arrays like 'line_items_0' as lists and keeps compound keys intact.
#     """
#     nested_dict = {}
#
#     for key, value in flattened_dict.items():
#         keys = key.split(sep)
#
#         d = nested_dict
#         for i, part in enumerate(keys):
#             # Check if part can be treated as an index for a list
#             if part.isdigit():
#                 part = int(part)
#
#             # Check if it's the last part of the key
#             if i == len(keys) - 1:
#                 d[part] = value
#             else:
#                 # Check the next part and handle list or dict accordingly
#                 if part not in d:
#                     if i + 1 < len(keys) and keys[i + 1].isdigit():
#                         d[part] = []
#                     else:
#                         d[part] = {}
#                 d = d[part]
#
#     return nested_dict
#
#
# def convert_dict_to_list(d):
#     """
#     Convert dictionaries with integer keys to lists to handle array-like structures.
#     """
#     if isinstance(d, dict):
#         if all(isinstance(k, int) for k in d.keys()):
#             # If all keys are integers, convert to list
#             return [convert_dict_to_list(v) for k, v in sorted(d.items())]
#         else:
#             # Otherwise, keep it as a dict
#             return {k: convert_dict_to_list(v) for k, v in d.items()}
#     elif isinstance(d, list):
#         return [convert_dict_to_list(v) for v in d]
#     else:
#         return d
#
#
# def excel_to_json(file_name, sheet_name, start_row=2):
#     """
#     Read Excel file and convert key-value pairs starting from a specific row to JSON format.
#     """
#     # Read the Excel sheet
#     df = pd.read_excel(file_name, sheet_name=sheet_name, header=None)
#
#     # Print the columns to debug
#     print("Columns in the Excel sheet:", df.columns)
#
#     # Skip rows until you reach the key-value pairs (starting from the provided start_row)
#     df = df.iloc[start_row:, :2]  # Adjust to take only key-value pairs (2 columns)
#     df.columns = ['Key', 'Value']  # Rename the columns after slicing
#
#     # Convert the dataframe to dictionary
#     flattened_dict = dict(zip(df['Key'], df['Value']))
#
#     # Convert the flattened dictionary back to nested structure
#     nested_data = unflatten_json(flattened_dict)
#
#     # Convert any dicts with integer keys to lists
#     nested_data = convert_dict_to_list(nested_data)
#
#     return nested_data
#
#
# # Example usage
# file_name = 'accountvault_data.xlsx'
# sheet_name = input("Enter sheet name: ")
#
# # Convert the Excel back to JSON
# nested_json = excel_to_json(file_name, sheet_name, start_row=2)  # Adjust 'start_row' if key-value pairs start later
#
# # Pretty-print the JSON output
# print(json.dumps(nested_json, indent=4))


# import pandas as pd
# from collections import defaultdict
# #a cross platform for clipboard platorm for python (now only for simple text)
# import pyperclip
#
# def flatten_key(key_parts):
#     """
#     Flatten the key parts into a single string with custom transformations.
#     """
#     # Custom transformations to match the desired output
#     if len(key_parts) == 2 and key_parts[0] == 'transaction' and key_parts[1] == 'id':
#         return 'transaction_id'
#
#     # Join the key parts and replace specific patterns
#     flat_key = '_'.join(key_parts)
#     flat_key = flat_key.replace('tax_exempt', 'tax_exempt')
#     flat_key = flat_key.replace('tax_amount', 'tax_amount')
#     flat_key = flat_key.replace('zip_code', 'zip_code')
#     flat_key = flat_key.replace('vat_registration', 'vat_registration')
#     flat_key = flat_key.replace('unique_vat_ref_number', 'unique_vat_ref_number')
#     return flat_key
#
#
# def unflatten_json(flattened_dict, sep='_'):
#     """
#     Convert flattened dictionary back to nested dictionary with custom key transformations.
#     """
#     nested_dict = defaultdict(dict)
#     for key, value in flattened_dict.items():
#         parts = key.split(sep)
#         d = nested_dict
#         for part in parts[:-1]:
#             if part.isdigit():
#                 part = int(part)
#             d = d.setdefault(part, {})
#         # Apply custom key flattening
#         flat_key = flatten_key(parts)
#         d[flat_key] = value
#     return dict(nested_dict)
#
#
# def excel_to_json(file_name, sheet_name, start_row=2):
#     """
#     Read Excel file and convert key-value pairs starting from a specific row to JSON format.
#     """
#     # Read the Excel sheet
#     df = pd.read_excel(file_name, sheet_name=sheet_name, header=None)
#
#     # Skip rows until you reach the key-value pairs (starting from the provided start_row)
#     df = df.iloc[start_row:, :2]  # Adjust to take only key-value pairs (2 columns)
#     df.columns = ['Key', 'Value']  # Rename the columns after slicing
#
#     # Convert the dataframe to dictionary
#     flattened_dict = dict(zip(df['Key'], df['Value']))
#
#     # Convert the flattened dictionary back to nested structure with custom formatting
#     nested_data = unflatten_json(flattened_dict)
#
#     return nested_data
#
#
# # Example usage
# file_name = 'accountvault_data.xlsx'
# sheet_name = input("Enter sheet name: ")
#
# # Convert the Excel back to JSON
# nested_json = excel_to_json(file_name, sheet_name, start_row=2)  # Adjust 'start_row' if key-value pairs start later
#
# print(nested_json)
# pyperclip.copy(nested_json)

# print("JSON data has been copied to the clipboard!")








#working nestedly

# import pandas as pd
# from collections import defaultdict
# import json
#
# def flatten_key(key_parts):
#     """
#     Flatten the key parts into a single string with custom transformations.
#     """
#     # Custom transformations to match the desired output
#     if len(key_parts) == 2 and key_parts[0] == 'transaction' and key_parts[1] == 'id':
#         return 'transaction_id'
#
#     # Join the key parts and replace specific patterns
#     flat_key = '_'.join(key_parts)
#     flat_key = flat_key.replace('tax_exempt', 'tax_exempt')
#     flat_key = flat_key.replace('tax_amount', 'tax_amount')
#     flat_key = flat_key.replace('zip_code', 'zip_code')
#     flat_key = flat_key.replace('vat_registration', 'vat_registration')
#     flat_key = flat_key.replace('unique_vat_ref_number', 'unique_vat_ref_number')
#     return flat_key
#
# def unflatten_json(flattened_dict, sep='_'):
#     """
#     Convert flattened dictionary back to nested dictionary with custom key transformations.
#     """
#     nested_dict = defaultdict(dict)
#     for key, value in flattened_dict.items():
#         parts = key.split(sep)
#         d = nested_dict
#         for part in parts[:-1]:
#             if part.isdigit():
#                 part = int(part)
#             d = d.setdefault(part, {})
#         # Apply custom key flattening
#         flat_key = flatten_key(parts)
#         d[flat_key] = value
#     return dict(nested_dict)
#
# def excel_to_json(file_name, sheet_name, start_row=2):
#     """
#     Read Excel file and convert key-value pairs starting from a specific row to JSON format.
#     """
#     # Read the Excel sheet
#     df = pd.read_excel(file_name, sheet_name=sheet_name, header=None)
#
#     # Skip rows until you reach the key-value pairs (starting from the provided start_row)
#     df = df.iloc[start_row:, :2]  # Adjust to take only key-value pairs (2 columns)
#     df.columns = ['Key', 'Value']  # Rename the columns after slicing
#
#     # Convert the dataframe to dictionary
#     flattened_dict = dict(zip(df['Key'], df['Value']))
#
#     # Convert the flattened dictionary back to nested structure with custom formatting
#     nested_data = unflatten_json(flattened_dict)
#
#     return nested_data
#
# # Example usage
# file_name = 'accountvault_data.xlsx'
# sheet_name = input("Enter sheet name: ")
#
# # Convert the Excel back to JSON
# nested_json = excel_to_json(file_name, sheet_name, start_row=2)  # Adjust 'start_row' if key-value pairs start later
#
# # Pretty-print the JSON
# pretty_json = json.dumps(nested_json, indent=4)
# print(pretty_json)


import pandas as pd
from collections import defaultdict
import json

from openpyxl import load_workbook
def flatten_key(key_parts):
    """
    Flatten the key parts into a single string with custom transformations.
    """
    if len(key_parts) == 2 and key_parts[0] == 'transaction' and key_parts[1] == 'id':
        return 'transaction_id'

    # Custom key flattening for line_items
    if key_parts[0] == "line_items" and key_parts[1].isdigit():
        return '_'.join(key_parts)

    # Join the key parts
    flat_key = '_'.join(key_parts)

    # Ensure correct mappings for known key structures
    flat_key = flat_key.replace('tax_exempt', 'tax_exempt')
    flat_key = flat_key.replace('sales_tax', 'sales_tax')
    flat_key = flat_key.replace('national_tax', 'national_tax')
    flat_key = flat_key.replace('vat_registration', 'vat_registration')
    flat_key = flat_key.replace('unique_vat_ref_number', 'unique_vat_ref_number')
    flat_key = flat_key.replace('zip_code', 'zip_code')

    return flat_key


def unflatten_json(file_name, sheet_name, flattened_dict, sep='_'):
    """
    Convert flattened dictionary back to nested dictionary with correct key structure.
    """
    nested_dict = defaultdict(dict)
    for key, value in flattened_dict.items():
        parts = key.split(sep)

        # Direct mapping for known root keys
        if parts[0] == "transactionlevel3":
            nested_dict["transactionlevel3"]["transaction_id"] = flattened_dict.get("transactionlevel3_transaction_id")
        else:
            # file_name = 'accountvault_data.xlsx'
            workbook = load_workbook(file_name)

            # Get the sheet by name (replace 'nested data' with your actual sheet name)
            sheet = workbook[sheet_name]

            # Find the first non-empty cell in the sheet
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:  # Checking for the first non-empty cell
                        first_word = str(cell)
                        break
                if cell is not None:
                    break

            # Print the first word found
            # print("First word in the sheet:", first_word)

            # Create level3_data structure
            # level3_data = nested_dict["transactionlevel333333"].setdefault("level3_dataaaaaaa", {})

            level3_data = nested_dict[first_word]

            # Special handling for line_items as a list
            if parts[0] == "line_items" and parts[1].isdigit():
                line_index = int(parts[1])
                line_items = level3_data.setdefault("line_items", [])
                # Extend the list if the index doesn't exist yet
                while len(line_items) <= line_index:
                    line_items.append({})
                item_key = '_'.join(parts[2:])
                line_items[line_index][item_key] = value
            else:
                flat_key = flatten_key(parts)
                level3_data[flat_key] = value

    return dict(nested_dict)


def excel_to_json(file_name, sheet_name, start_row=2):
    """
    Read Excel file and convert key-value pairs starting from a specific row to JSON format.
    """
    # Read the Excel sheet
    df = pd.read_excel(file_name, sheet_name=sheet_name, header=None)

    # Skip rows until you reach the key-value pairs (starting from the provided start_row)
    df = df.iloc[start_row:, :2]  # Adjust to take only key-value pairs (2 columns)
    df.columns = ['Key', 'Value']  # Rename the columns after slicing

    # Convert the dataframe to dictionary
    flattened_dict = dict(zip(df['Key'], df['Value']))

    # Convert the flattened dictionary back to nested structure with custom formatting
    nested_data = unflatten_json(file_name, sheet_name, flattened_dict)

    return nested_data


# Example usage
file_name = 'accountvault_data.xlsx'
sheet_name = input("Enter sheet name: ")

# Convert the Excel back to JSON
nested_json = excel_to_json(file_name, sheet_name, start_row=2)  # Adjust 'start_row' if key-value pairs start later

# Pretty-print the JSON
pretty_json = json.dumps(nested_json, indent=4)
print(pretty_json)
