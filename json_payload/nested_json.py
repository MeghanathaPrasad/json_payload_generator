import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def flatten_json(nested_json, parent_key='', sep='.'):
    items = []
    for k, v in nested_json.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_json(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)

def add_sheet_to_excel(file_name, sheet_name, data, table_name):
    # Flatten the JSON to handle nested structures
    flattened_data = flatten_json(data)
    data_df = pd.DataFrame(list(flattened_data.items()), columns=["Key", "Value"])
    table_name_df = pd.DataFrame([[table_name]])

    if os.path.exists(file_name):
        try:
            # Load the existing workbook
            book = load_workbook(file_name)

            # Check if the sheet already exists
            if sheet_name in book.sheetnames:
                # Append to the existing sheet
                sheet = book[sheet_name]

                # Write table name and data at the next empty row
                next_row = sheet.max_row + 1
                for r in dataframe_to_rows(table_name_df, index=False, header=False):
                    sheet.append(r)
                for r in dataframe_to_rows(data_df, index=False, header=False):
                    sheet.append(r)

            else:
                # Create a new sheet
                sheet = book.create_sheet(title=sheet_name)

                # Write table name and data
                for r in dataframe_to_rows(table_name_df, index=False, header=False):
                    sheet.append(r)
                for r in dataframe_to_rows(data_df, index=False, header=True):
                    sheet.append(r)

            # Save the workbook
            book.save(file_name)
            print(f"Data successfully added to '{sheet_name}' in {file_name}.")

        except Exception as e:
            print(f"Error processing the file: {e}.")
    else:
        # Create a new file and add the first sheet
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            # Create the DataFrame and write the table name and data
            table_name_df.to_excel(writer, sheet_name=sheet_name, header=False, index=False, startrow=0)
            data_df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)

        print(f"File created and data written to '{sheet_name}'.")

# Example usage
data ={
        "transaction_id": "111111111111111111111111",
        "level3_data":{
            "tax_exempt": "0",
            "sales_tax":"200",
            "national_tax":"2",
            "merchant_vat_registration": "123456",
            "customer_vat_registration": "12345678",
            "summary_commodity_code": "C1K2",
            "freight_amount": "0.1",
            "duty_amount": "0",
            "shipto_zip_code": "FL1234",
            "shipfrom_zip_code": "AZ1234",
            "destination_country_code": "840",
            "unique_vat_ref_number": "vat1234",
            "order_date": "171006",
            "tax_amount": "0",
            "tax_rate": "0",
            "line_items": [
                {
                    "commodity_code": "cc123456",
                    "description": "cool drink",
                    "product_code": "coke12678",
                    "quantity": "5",
                    "unit_code": "gll",
                    "unit_cost": "4",
                    "tax_amount": "10",
                    "tax_rate": "0",
                    "discount_amount": "0",
                    "other_tax_amount": "0"
                },
                {
                    "commodity_code": "cc1234",
                    "description": "cool drink",
                    "product_code": "fanta123678",
                    "quantity": "12",
                    "unit_code": "gll",
                    "unit_cost": "3",
                    "tax_amount": "4",
                    "tax_rate": "0",
                    "discount_amount": "7",
                    "other_tax_amount": "0"
                }
            ]
        }
    }
file_name = 'accountvault_data.xlsx'
sheet_name = input("Enter sheet name: ")
table_name = input("Enter table name: ")

# Add sheet to the existing Excel file
add_sheet_to_excel(file_name, sheet_name, data, table_name)


# import pandas as pd
# from collections import defaultdict
#
#
# def flatten_key(key_parts):
#     """
#     Flatten the key parts into a single string with custom transformations.
#     """
#     flat_key = '_'.join(key_parts)
#
#     # Custom transformations to match the desired output
#     # Handle special cases directly
#     if flat_key == 'transaction_id':
#         return 'transaction_id'
#     if flat_key.startswith('level3_data'):
#         flat_key = flat_key.replace('level3_data_', '')
#     if flat_key.startswith('transaction'):
#         flat_key = flat_key.replace('transaction_', 'transactionlevel3_')
#
#     # Return the flattened key after transformations
#     return flat_key
#
#
# def unflatten_json(flattened_dict, sep='_'):
#     """
#     Convert flattened dictionary back to nested dictionary with custom key transformations.
#     """
#     nested_dict = defaultdict(dict)
#     for key, value in flattened_dict.items():
#         parts = key.split(sep)
#         d = nested_dict
#         for part in parts[:-1]:
#             if part.isdigit():
#                 part = int(part)
#             d = d.setdefault(part, {})
#         # Apply custom key flattening
#         flat_key = flatten_key(parts)
#         d[flat_key] = value
#     return dict(nested_dict)
#
#
# def excel_to_json(file_name, sheet_name, start_row=2):
#     """
#     Read Excel file and convert key-value pairs starting from a specific row to JSON format.
#     """
#     # Read the Excel sheet
#     df = pd.read_excel(file_name, sheet_name=sheet_name, header=None)
#
#     # Skip rows until you reach the key-value pairs (starting from the provided start_row)
#     df = df.iloc[start_row:, :2]  # Adjust to take only key-value pairs (2 columns)
#     df.columns = ['Key', 'Value']  # Rename the columns after slicing
#
#     # Convert the dataframe to dictionary
#     flattened_dict = dict(zip(df['Key'], df['Value']))
#
#     # Convert the flattened dictionary back to nested structure with custom formatting
#     nested_data = unflatten_json(flattened_dict)
#
#     return nested_data
#
#
# # Example usage
# file_name = 'accountvault_data.xlsx'
# sheet_name = input("Enter sheet name: ")
#
# # Convert the Excel back to JSON
# nested_json = excel_to_json(file_name, sheet_name, start_row=2)  # Adjust 'start_row' if key-value pairs start later
#
# print(nested_json)
#
