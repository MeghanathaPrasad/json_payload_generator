import json
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


with open('data.json', 'r') as file:
    nested_dict = json.load(file)

def flatten_dict(d, parent_key='', sep='.'):
    """
    Recursively flattens a nested dictionary and handles lists, tuples, and other data types.

    Args:
    - d (dict): The nested dictionary to flatten.
    - parent_key (str): The base key string for the flattened keys.
    - sep (str): The separator between keys.

    Returns:
    - dict: The flattened dictionary.
    """
    flat_dict = {}
    for k, v in d.items():
        new_key = f'{parent_key}{sep}{k}' if parent_key else k  # Construct new key

        if isinstance(v, dict):  # Handle nested dictionary
            flat_dict.update(flatten_dict(v, new_key, sep=sep))

        elif isinstance(v, (list, tuple)):  # Handle lists or tuples
            for i, item in enumerate(v):
                flat_dict.update(flatten_dict({f"{new_key}{sep}{i}": item}, '', sep=sep))

        else:  # Handle other types (str, int, float, etc.)
            flat_dict[new_key] = v

    return flat_dict




# df.to_excel("excel_output.xlsx", index=False)


def add_sheet_to_excel(file_name, sheet_name, df):
    # data_df = pd.DataFrame([data])
    # table_name_df = pd.DataFrame([[table_name]])

    if os.path.exists(file_name):
        try:
            # Load the existing workbook
            book = load_workbook(file_name)

            # Check if the sheet already exists
            if sheet_name in book.sheetnames:
                # Append to the existing sheet
                sheet = book[sheet_name]

                # Find the next empty row
                next_row = sheet.max_row + 1

                # Write table name and data at the next empty row
                # for r in dataframe_to_rows(index=False, header=False):
                #     sheet.append(r)
                for r in dataframe_to_rows(df, index=False, header=False):
                    sheet.append(r)

            else:
                # Create a new sheet
                sheet = book.create_sheet(title=sheet_name)


                # Write table name and data
                # for r in dataframe_to_rows(table_name_df, index=False, header=False):
                #     sheet.append(r)
                for r in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(r)




            # Save the workbook
            book.save(file_name)
            print(f"Data successfully added to '{sheet_name}' in {file_name}.")

        except Exception as e:
            print(f"Error processing the file: {e}.")
    else:
        # Create a new file and add the first sheet
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            # Create the DataFrame
            df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)

        print(f"File created and data written to '{sheet_name}'.")

# Flatten the nested dictionary
flattened_dict = flatten_dict(nested_dict)

# print(flattened_dict)flattened_dict

# Convert flattened dictionary to a list of key-value pairs
data = list(flattened_dict.items())

# Convert to DataFrame
df = pd.DataFrame(data, columns=['Key', 'Value'])

print(df)

file_name = 'excel_output.xlsx'
sheet_name = input("Enter sheet name: ")

#add the flattened_dict in excel
add_sheet_to_excel(file_name,sheet_name,df)
