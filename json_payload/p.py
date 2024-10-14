
# import pandas as pd
# from collections import defaultdict
# import json
# from openpyxl import load_workbook
#
#
# def unflatten_json(file_name, sheet_name, flattened_dict, sep='_'):
#     """
#     Convert flattened dictionary back to nested dictionary with correct key structure.
#     """
#     nested_dict = defaultdict(dict)
#     for key, value in flattened_dict.items():
#         # parts = key.split(sep)
#
#         workbook = load_workbook(file_name)
#
#         # Get the sheet by name (replace 'nested data' with your actual sheet name)
#         sheet = workbook[sheet_name]
#
#         # Find the first non-empty cell in the sheet
#         for row in sheet.iter_rows(values_only=True):
#             for cell in row:
#                 if cell is not None:  # Checking for the first non-empty cell
#                     first_word = str(cell)
#                     break
#             if cell is not None:
#                 break
#
#
#         level3_data = nested_dict[first_word]
#         level3_data[key] = value
#
#     return dict(nested_dict)
#
#
# def excel_to_json(file_name, sheet_name, start_row=2):
#     """
#     Read Excel file and convert key-value pairs starting from a specific row to JSON format.
#     """
#     # Read the Excel sheet
#     df = pd.read_excel(file_name, sheet_name=sheet_name, header=None)
#
#     # Skip rows until you reach the key-value pairs (starting from the provided start_row)
#     df = df.iloc[start_row:, :2]  # Adjust to take only key-value pairs (2 columns)
#     df.columns = ['Key', 'Value']  # Rename the columns after slicing
#
#     # Convert the dataframe to dictionary
#     flattened_dict = dict(zip(df['Key'], df['Value']))
#
#     # Convert the flattened dictionary back to nested structure with custom formatting
#     nested_data = unflatten_json(file_name, sheet_name, flattened_dict)
#
#     return nested_data
#
#
# # Example usage
# file_name = 'accountvault_data.xlsx'
# sheet_name = input("Enter sheet name: ")
#
# # Convert the Excel back to JSON
# nested_json = excel_to_json(file_name, sheet_name, start_row=2)  # Adjust 'start_row' if key-value pairs start later
#
# # Pretty-print the JSON
# pretty_json = json.dumps(nested_json, indent=4)
# print(pretty_json)




import pandas as pd
from collections import defaultdict
import json
from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox


def unflatten_json(file_name, sheet_name, flattened_dict, sep='_'):
    """
    Convert flattened dictionary back to nested dictionary with correct key structure.
    """
    nested_dict = defaultdict(dict)
    for key, value in flattened_dict.items():
        # parts = key.split(sep)

        workbook = load_workbook(file_name)

        # Get the sheet by name (replace 'nested data' with your actual sheet name)
        sheet = workbook[sheet_name]

        # Find the first non-empty cell in the sheet
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:  # Checking for the first non-empty cell
                    first_word = str(cell)
                    break
            if cell is not None:
                break

        level3_data = nested_dict[first_word]
        level3_data[key] = value

    return dict(nested_dict)


def excel_to_json(file_name, sheet_name, start_row=2):
    """
    Read Excel file and convert key-value pairs starting from a specific row to JSON format.
    """
    # Read the Excel sheet
    df = pd.read_excel(file_name, sheet_name=sheet_name, header=None)

    # Skip rows until you reach the key-value pairs (starting from the provided start_row)
    df = df.iloc[start_row:, :2]  # Adjust to take only key-value pairs (2 columns)
    df.columns = ['Key', 'Value']  # Rename the columns after slicing

    # Convert the dataframe to dictionary
    flattened_dict = dict(zip(df['Key'], df['Value']))

    # Convert the flattened dictionary back to nested structure with custom formatting
    nested_data = unflatten_json(file_name, sheet_name, flattened_dict)

    return nested_data


def copy_to_clipboard(json_data):
    """
    Copies the JSON data to the clipboard.
    """
    root = tk.Tk()
    root.withdraw()  # Hide the main tkinter window
    root.clipboard_clear()
    root.clipboard_append(json_data)
    root.update()  # Keep the clipboard data available after the window is closed
    root.destroy()
    messagebox.showinfo("Copied", "JSON data has been copied to clipboard")


def show_json_output(json_data):
    """
    Displays the JSON data in a tkinter window with a Copy button.
    """
    root = tk.Tk()
    root.title("JSON Output")

    # Create a text widget to display the JSON data
    text_widget = tk.Text(root, wrap="word", height=30, width=80)
    text_widget.insert("1.0", json_data)
    text_widget.pack()

    # Create a copy button
    copy_button = tk.Button(root, text="Copy to Clipboard", command=lambda: copy_to_clipboard(json_data))
    copy_button.pack()

    root.mainloop()


# Example usage
file_name = 'accountvault_data.xlsx'
sheet_name = input("Enter sheet name: ")

# Convert the Excel back to JSON
nested_json = excel_to_json(file_name, sheet_name, start_row=2)  # Adjust 'start_row' if key-value pairs start later

# Pretty-print the JSON
pretty_json = json.dumps(nested_json, indent=4)

# Display the JSON output with a copy button
show_json_output(pretty_json)
